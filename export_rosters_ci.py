#!/usr/bin/env python3
"""
export_rosters_ci.py — Cloud version of export_rosters.py.
Reads all_draftboards_complete.xlsx from the repo root
and writes rosters.json. Run via GitHub Actions workflow_dispatch.
"""

import json
import unicodedata
from pathlib import Path
from openpyxl import load_workbook

XLSX_PATH = Path("all_draftboards_complete.xlsx")

def strip_accents(name):
    return "".join(c for c in unicodedata.normalize("NFD", str(name))
                   if unicodedata.category(c) != "Mn")

def main():
    if not XLSX_PATH.exists():
        raise FileNotFoundError(
            f"{XLSX_PATH} not found in repo root. "
            "Make sure you've pushed all_draftboards_complete.xlsx to GitHub."
        )

    print(f"Loading {XLSX_PATH}...")
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)

    rosters = {}
    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("draftboard_"):
            continue
        ws = wb[sheet_name]
        players = []
        for row in range(2, 242):
            name = ws.cell(row, 2).value
            team = ws.cell(row, 5).value
            if not name or not team:
                continue
            players.append({
                "pick":      ws.cell(row, 1).value,
                "name":      strip_accents(str(name).strip()),
                "pos":       str(ws.cell(row, 3).value).strip(),
                "mlb":       str(ws.cell(row, 4).value).strip().upper().replace("AZ","ARI"),
                "team_name": str(team).strip(),
            })
        rosters[sheet_name] = players
        print(f"  {sheet_name}: {len(players)} players")

    out = Path("rosters.json")
    out.write_text(json.dumps(rosters, indent=2))
    print(f"\n✓ rosters.json saved ({out.stat().st_size:,} bytes)")
    print(f"  {len(rosters)} draftboards, "
          f"{sum(len(v) for v in rosters.values())} total player slots")

if __name__ == "__main__":
    main()
