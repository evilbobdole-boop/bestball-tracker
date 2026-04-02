#!/usr/bin/env python3
"""
MLB Best Ball $600K Slugfest — Full Leaderboard Publisher

Reads rosters + stats from the xlsx, computes scores,
generates index.html with season summary + all 35 draft pages,
then pushes to GitHub Pages.

Usage:
    python publish.py                   # generate + push
    python publish.py --local           # generate only, no push
    python publish.py --weeks 2         # score through week 2

One-time setup:
    1. Create a PUBLIC GitHub repo e.g. "bestball-tracker"
    2. Repo Settings -> Pages -> Deploy from branch -> main -> / (root)
    3. Clone repo locally, put this script inside it
    4. Edit XLSX_PATH and REPO_DIR below
    5. Run: python publish.py
    6. Live at: https://YOUR_USERNAME.github.io/bestball-tracker/
"""

import argparse
import subprocess
import unicodedata
from datetime import datetime
from pathlib import Path
from openpyxl import load_workbook

# ── EDIT THESE TWO PATHS ─────────────────────────────────────────────────────
XLSX_PATH = Path(r"C:\Users\EvilBobFUCKINGDole\Downloads\BEST BALL 2026\all_draftboards_complete.xlsx")
REPO_DIR  = Path(r"C:\Users\EvilBobFUCKINGDole\Downloads\BEST BALL 2026\bestball-tracker")
# ─────────────────────────────────────────────────────────────────────────────

MY_TEAM = "evilbobdole"


# ── HELPERS ───────────────────────────────────────────────────────────────────

def strip_accents(name: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", str(name))
        if unicodedata.category(c) != "Mn"
    )

def ordinal(n):
    if n is None: return "—"
    s = ["th","st","nd","rd"] + ["th"] * 16
    return f"{n}{s[n % 100 if 11 <= n % 100 <= 13 else n % 10]}"

def load_stats(wb):
    batting        = {}   # (week, name, team) -> total dk pts
    pitching       = {}   # (week, name, team) -> total dk pts
    batting_daily  = {}   # (date_str, name, team) -> dk pts
    pitching_daily = {}   # (date_str, name, team) -> dk pts

    ds = wb["Daily Stats"]
    # Use sheet max_row but also read all rows regardless of gaps
    for row in ds.iter_rows(min_row=2, max_row=ds.max_row, values_only=True):
        # Skip completely empty rows or the instruction placeholder row
        if row[2] is None or row[1] is None: continue
        name_str = str(row[2]).strip()
        if not name_str or name_str.startswith("←"): continue
        try:
            week = int(float(row[1]))
            name = strip_accents(name_str)
            team = str(row[3]).strip().upper().replace("AZ","ARI")
            dk   = float(row[18]) if row[18] is not None else 0.0
            if week < 1: continue
            key  = (week, name, team)
            batting[key] = batting.get(key, 0.0) + dk
            # Daily tracking
            date_str = str(row[0])[:10]
            dkey = (date_str, name, team)
            batting_daily[dkey] = batting_daily.get(dkey, 0.0) + dk
        except (TypeError, ValueError):
            continue

    ps = wb["Pitching Stats"]
    for row in ps.iter_rows(min_row=2, max_row=ps.max_row, values_only=True):
        if row[2] is None or row[1] is None: continue
        name_str = str(row[2]).strip()
        if not name_str or name_str.startswith("←"): continue
        try:
            week = int(float(row[1]))
            name = strip_accents(name_str)
            team = str(row[3]).strip().upper().replace("AZ","ARI")
            dk   = float(row[11]) if row[11] is not None else 0.0
            if week < 1: continue
            key  = (week, name, team)
            pitching[key] = pitching.get(key, 0.0) + dk
            date_str = str(row[0])[:10]
            dkey = (date_str, name, team)
            pitching_daily[dkey] = pitching_daily.get(dkey, 0.0) + dk
        except (TypeError, ValueError):
            continue

    return batting, pitching, batting_daily, pitching_daily

def player_score(name, mlb, pos, week, batting, pitching):
    name = strip_accents(str(name).strip())
    team = str(mlb).strip().upper().replace("AZ","ARI")
    if pos == "P":
        return pitching.get((week, name, team), 0.0)
    return batting.get((week, name, team), 0.0)

def team_week_score(roster, team_name, week, batting, pitching):
    by_pos = {"P": [], "IF": [], "OF": []}
    for p in roster:
        if p["team_name"] != team_name or p["pos"] not in by_pos: continue
        s = player_score(p["name"], p["mlb"], p["pos"], week, batting, pitching)
        by_pos[p["pos"]].append(s)
    total = 0.0
    for scores in by_pos.values():
        scores.sort(reverse=True)
        total += sum(scores[:3])
    return round(total, 2)


def team_day_score(roster, team_name, date_str, batting_daily, pitching_daily,
                   batting, pitching, week):
    """Daily score for a team's STARTERS only.
    Starters = top 3 per position by WEEK score (not today's score).
    Only those 9 players' daily points count — bench players are excluded.
    """
    by_pos = {"P": [], "IF": [], "OF": []}
    # Build per-player week totals to identify starters
    for p in roster:
        if p["team_name"] != team_name or p["pos"] not in by_pos: continue
        week_pts = player_score(p["name"], p["mlb"], p["pos"], week, batting, pitching)
        by_pos[p["pos"]].append((p, week_pts))

    total = 0.0
    for pos, players in by_pos.items():
        # Sort by week score to find starters
        starters = sorted(players, key=lambda x: x[1], reverse=True)[:3]
        for p, _ in starters:
            name = strip_accents(str(p["name"]).strip())
            team = str(p["mlb"]).strip().upper().replace("AZ","ARI")
            key  = (date_str, name, team)
            if p["pos"] == "P":
                total += pitching_daily.get(key, 0.0)
            else:
                total += batting_daily.get(key, 0.0)
    return round(total, 2)

def load_all_drafts(wb, batting, pitching, batting_daily, pitching_daily, latest_date, yesterday_date, num_weeks, adp_players=None):
    drafts = []

    def top3_by_pos(players):
        """Top 3 per position by week score — the starters."""
        by_pos = {"IF": [], "OF": [], "P": []}
        for p in players:
            if p["pos"] in by_pos:
                by_pos[p["pos"]].append(p)
        result = []
        for pos in ["P", "IF", "OF"]:
            sorted_pos = sorted(by_pos[pos], key=lambda x: x["weeks"][-1] if x["weeks"] else 0, reverse=True)
            result.extend(sorted_pos[:3])
        return result

    def build_players(team_name):
        """Build per-player week + daily + yesterday scores, filtered to week starters."""
        players = []
        for p in roster:
            if p["team_name"] != team_name: continue
            weeks  = [round(player_score(p["name"], p["mlb"], p["pos"], w, batting, pitching), 2)
                      for w in range(1, num_weeks + 1)]
            name_s = strip_accents(str(p["name"]).strip())
            team_s = str(p["mlb"]).strip().upper().replace("AZ","ARI")
            src    = pitching_daily if p["pos"] == "P" else batting_daily
            daily  = round(src.get((latest_date,    name_s, team_s), 0.0), 2) if latest_date    else 0.0
            yest   = round(src.get((yesterday_date, name_s, team_s), 0.0), 2) if yesterday_date else 0.0
            players.append({**p, "weeks": weeks, "total": round(sum(weeks), 2), "daily": daily, "yesterday": yest})
        return top3_by_pos(players)

    def build_bench(team_name):
        """Bench = all players NOT in top 3 per pos by week score.
        Sorted P/IF/OF then week total high to low.
        Includes Yest/Today for reference only — not part of team total."""
        all_players = []
        for p in roster:
            if p["team_name"] != team_name: continue
            weeks  = [round(player_score(p["name"], p["mlb"], p["pos"], w, batting, pitching), 2)
                      for w in range(1, num_weeks + 1)]
            name_s = strip_accents(str(p["name"]).strip())
            team_s = str(p["mlb"]).strip().upper().replace("AZ","ARI")
            src_d  = pitching_daily if p["pos"] == "P" else batting_daily
            src_y  = pitching_daily if p["pos"] == "P" else batting_daily
            # Use yesterday_date key for yest
            daily  = round((pitching_daily if p["pos"]=="P" else batting_daily).get((latest_date,    name_s, team_s), 0.0), 2) if latest_date    else 0.0
            yest   = round((pitching_daily if p["pos"]=="P" else batting_daily).get((yesterday_date, name_s, team_s), 0.0), 2) if yesterday_date else 0.0
            # Fix: use pitching_yest/batting_yest equivalent from daily dicts
            # batting_daily only has latest_date entries — need full batting for yesterday
            yest   = 0.0  # will be filled below
            if yesterday_date:
                yest_src = pitching_daily if p["pos"] == "P" else batting_daily
                # Actually batting_daily is keyed by latest_date only
                # Use the raw batting/pitching dicts filtered by yesterday_date
                from_raw = pitching if p["pos"] == "P" else batting
                yest = round(sum(v for k, v in from_raw.items()
                                 if k[0] == yesterday_date and k[1] == name_s and k[2] == team_s), 2)
            daily_raw = pitching if p["pos"] == "P" else batting
            daily  = round(sum(v for k, v in daily_raw.items()
                               if k[0] == latest_date and k[1] == name_s and k[2] == team_s), 2) if latest_date else 0.0
            all_players.append({**p, "weeks": weeks, "total": round(sum(weeks), 2),
                                 "daily": daily, "yesterday": yest})

        # Find starters (top 3 per pos)
        starters = set()
        by_pos = {"P": [], "IF": [], "OF": []}
        for p in all_players:
            if p["pos"] in by_pos:
                by_pos[p["pos"]].append(p)
        for pos, plist in by_pos.items():
            for p in sorted(plist, key=lambda x: x["total"], reverse=True)[:3]:
                starters.add(p["name"])

        # Bench = everyone not a starter, sorted P/IF/OF then total desc
        pos_order = {"P": 0, "IF": 1, "OF": 2}
        bench = [p for p in all_players if p["name"] not in starters]
        bench.sort(key=lambda x: (pos_order.get(x["pos"], 9), -x["total"]))
        return bench

    for sheet_name in wb.sheetnames:
        if not sheet_name.startswith("draftboard_"): continue
        ws  = wb[sheet_name]
        num = int(sheet_name.replace("draftboard_","").replace("_",""))

        # Roster
        roster = []
        for row in range(2, 242):
            name = ws.cell(row, 2).value
            team = ws.cell(row, 5).value
            if not name or not team: continue
            roster.append({
                "pick":      ws.cell(row, 1).value,
                "name":      str(name).strip(),
                "pos":       str(ws.cell(row, 3).value).strip(),
                "mlb":       str(ws.cell(row, 4).value).strip(),
                "team_name": str(team).strip(),
            })

        # Teams in draft order from leaderboard col G
        teams = []
        seen  = set()
        for row in range(2, 14):
            t = ws.cell(row, 7).value
            if t and str(t).strip() not in seen:
                teams.append(str(t).strip())
                seen.add(str(t).strip())

        # Season scores
        team_data = {}
        for team in teams:
            weekly = [team_week_score(roster, team, w, batting, pitching)
                      for w in range(1, num_weeks + 1)]
            team_data[team] = {"total": round(sum(weekly), 2), "weeks": weekly}

        ranked  = sorted(teams, key=lambda t: team_data[t]["total"], reverse=True)
        my_rank = ranked.index(MY_TEAM) + 1 if MY_TEAM in ranked else None
        my_pts  = team_data.get(MY_TEAM, {}).get("total", 0.0)

        # 2nd place team (by season score)
        second_team = ranked[1] if len(ranked) > 1 else None

        # Build starter player lists (top 3 per pos by week score)
        my_players     = build_players(MY_TEAM)     if MY_TEAM in teams else []
        second_players = build_players(second_team) if second_team       else []

        def starters_score_for_date(team_name, date_str):
            """Sum a date's score for the 9 week starters (top 3 per pos by week total)."""
            by_pos = {"P": [], "IF": [], "OF": []}
            for p in roster:
                if p["team_name"] != team_name or p["pos"] not in by_pos: continue
                week_pts = player_score(p["name"], p["mlb"], p["pos"], num_weeks, batting, pitching)
                by_pos[p["pos"]].append((p, week_pts))
            total = 0.0
            for pos, players in by_pos.items():
                starters = sorted(players, key=lambda x: x[1], reverse=True)[:3]
                for p, _ in starters:
                    if not date_str: continue
                    name_s = strip_accents(str(p["name"]).strip())
                    team_s = str(p["mlb"]).strip().upper().replace("AZ","ARI")
                    key    = (date_str, name_s, team_s)
                    total += pitching_daily.get(key, 0.0) if p["pos"] == "P" else batting_daily.get(key, 0.0)
            return round(total, 2)

        # Today
        my_daily      = starters_score_for_date(MY_TEAM,     latest_date)    if MY_TEAM in teams else 0.0
        second_today  = starters_score_for_date(second_team, latest_date)    if second_team      else 0.0
        my_daily_gap  = round(my_daily - second_today, 2)

        # Yesterday
        my_yesterday      = starters_score_for_date(MY_TEAM,     yesterday_date) if MY_TEAM in teams else 0.0
        second_yesterday  = starters_score_for_date(second_team, yesterday_date) if second_team      else 0.0
        my_yesterday_gap  = round(my_yesterday - second_yesterday, 2)

        # Daily scores for all teams for ranking
        daily_scores = {team: starters_score_for_date(team, latest_date) for team in teams} if latest_date else {}
        daily_ranked = sorted(teams, key=lambda t: daily_scores.get(t, 0.0), reverse=True)

        my_bench     = build_bench(MY_TEAM)     if MY_TEAM in teams else []
        second_bench = build_bench(second_team) if second_team       else []

        drafts.append({
            "num":              num,
            "sheet":            sheet_name,
            "roster":           roster,
            "teams":            teams,
            "data":             team_data,
            "ranked":           ranked,
            "my_rank":          my_rank,
            "my_pts":           my_pts,
            "my_players":       my_players,
            "my_bench":         my_bench,
            "second_team":      second_team,
            "second_players":   second_players,
            "second_bench":     second_bench,
            "daily_scores":     daily_scores,
            "my_daily":         my_daily,
            "second_today":     second_today,
            "my_daily_gap":     my_daily_gap,
            "my_yesterday":     my_yesterday,
            "second_yesterday": second_yesterday,
            "my_yesterday_gap": my_yesterday_gap,
            "latest_date":      latest_date,
            "yesterday_date":   yesterday_date,
        })

    drafts.sort(key=lambda d: d["num"])

    # ── Player analytics ──
    player_analytics = []
    if adp_players:
        total_drafts = len(drafts)
        for p in adp_players:
            pname = p["name"]
            ppos  = p["pos"]
            pmlb  = p["mlb"]

            # Count appearances across all drafts
            times_drafted    = 0  # by anyone
            drafted_by_me    = 0  # by evilbobdole
            cashing          = 0  # evilbobdole has them + draft top 2

            for d in drafts:
                # Is this player in this draft at all?
                in_draft   = any(r["name"] == pname and r["mlb"] == pmlb for r in d["roster"])
                on_my_team = any(r["name"] == pname and r["mlb"] == pmlb and r["team_name"] == MY_TEAM
                                 for r in d["roster"])
                if in_draft:
                    times_drafted += 1
                if on_my_team:
                    drafted_by_me += 1
                # Cashing = player is owned by ANY team currently in top 2
                if in_draft:
                    top2_teams = set(d["ranked"][:2]) if len(d["ranked"]) >= 2 else set(d["ranked"])
                    owner = next((r["team_name"] for r in d["roster"]
                                  if r["name"] == pname and r["mlb"] == pmlb), None)
                    if owner and owner in top2_teams:
                        cashing += 1

            draft_pct    = round(times_drafted / total_drafts * 100, 1)
            season_total = round(sum(
                player_score(pname, pmlb, ppos, w, batting, pitching)
                for w in range(1, num_weeks + 1)
            ), 2)
            week_totals  = [round(player_score(pname, pmlb, ppos, w, batting, pitching), 2)
                            for w in range(num_weeks, 0, -1)]  # latest first

            player_analytics.append({
                "name":         pname,
                "pos":          ppos,
                "mlb":          pmlb,
                "draft_pct":    draft_pct,
                "drafted_by_me":drafted_by_me,
                "cashing":      cashing,
                "season_total": season_total,
                "week_totals":  week_totals,
            })

    return drafts, player_analytics


# ── HTML ──────────────────────────────────────────────────────────────────────

CSS = """
:root {
  --navy:   #0D2D5E;
  --blue:   #1F4E79;
  --blue2:  #2E75B6;
  --gold:   #FFD700;
  --silver: #C0C0C0;
  --bronze: #CD7F32;
  --green:  #00CC00;
  --lgreen: #E6FFE6;
  --bg:     #F2F5FA;
  --card:   #FFFFFF;
  --text:   #222222;
  --muted:  #666666;
}
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: Arial, sans-serif; background: var(--bg); color: var(--text); }

/* NAV */
nav {
  background: var(--navy);
  padding: 0 16px;
  display: flex;
  align-items: center;
  gap: 6px;
  height: 46px;
  position: sticky;
  top: 0;
  z-index: 200;
  overflow-x: auto;
  white-space: nowrap;
}
nav .brand { color: #fff; font-weight: bold; font-size: 14px; margin-right: 10px; flex-shrink: 0; }
nav a { color: #aac; text-decoration: none; font-size: 11px; padding: 3px 7px;
        border-radius: 4px; flex-shrink: 0; }
nav a:hover { background: rgba(255,255,255,0.15); color: #fff; }

/* HERO */
.hero {
  background: linear-gradient(135deg, var(--navy) 0%, var(--blue2) 100%);
  color: #fff; padding: 36px 20px 28px; text-align: center;
}
.hero h1 { font-size: 26px; margin-bottom: 6px; }
.hero .sub { opacity: .75; font-size: 13px; }
.hero .ts  { opacity: .55; font-size: 11px; margin-top: 5px; }

/* STAT PILLS */
.pills { display: flex; justify-content: center; gap: 16px; flex-wrap: wrap;
         margin: 20px 16px 0; }
.pill { background: var(--card); border-radius: 10px; padding: 14px 20px;
        text-align: center; min-width: 110px; box-shadow: 0 2px 8px rgba(0,0,0,.08); }
.pill .val { font-size: 22px; font-weight: bold; color: var(--blue); }
.pill .lbl { font-size: 11px; color: var(--muted); margin-top: 2px; }

/* SECTION WRAPPER */
.container { max-width: 960px; margin: 0 auto; padding: 0 14px 60px; }

/* SEASON SUMMARY */
.summary {
  background: var(--card); border-radius: 10px;
  box-shadow: 0 2px 10px rgba(0,0,0,.08); margin: 24px 0; overflow: hidden;
}
.summary .sec-title {
  background: var(--blue); color: #fff; padding: 11px 18px; font-size: 15px; font-weight: bold;
}

/* DRAFT CARD */
.draft-card {
  background: var(--card); border-radius: 10px;
  box-shadow: 0 2px 8px rgba(0,0,0,.07); margin-bottom: 24px; overflow: hidden;
}
.draft-header {
  background: var(--blue); color: #fff; padding: 11px 18px;
  display: flex; align-items: center; justify-content: space-between; flex-wrap: wrap; gap: 8px;
}
.draft-header h2 { font-size: 15px; }
.badge {
  background: var(--green); color: #000; padding: 3px 11px;
  border-radius: 12px; font-size: 12px; font-weight: bold;
}

/* TABLES */
table { width: 100%; border-collapse: collapse; font-size: 13px; }
th { background: #E8EEF7; padding: 8px 12px; text-align: left; font-weight: bold; color: #333; }
td { padding: 7px 12px; border-bottom: 1px solid #F0F0F0; }
tr:last-child td { border-bottom: none; }
tr:hover td { background: #F7F9FC; }
.num  { font-family: monospace; }
.bold { font-weight: bold; }

/* RANK BADGES */
.r1 { background: var(--gold)   !important; color: #333; font-weight: bold; }
.r2 { background: var(--silver) !important; color: #333; font-weight: bold; }
.r3 { background: var(--bronze) !important; color: #fff; font-weight: bold; }
.rank-cell { text-align: center; width: 36px; font-weight: bold; }

/* MY TEAM ROW */
.my-row td { background: var(--lgreen) !important; font-weight: bold; }
.my-row:hover td { background: #d0f5d0 !important; }

/* PLAYER SECTION */
.player-section { border-top: 3px solid var(--green); }
.player-header  { background: var(--green); color: #000; padding: 8px 16px; font-size: 13px; font-weight: bold; }
.pos-IF td { background: #FFF0E8; }
.pos-OF td { background: #F0F8EE; }
.pos-P  td { background: #EAF2FB; }
.pos-break td { border-top: 2px solid #aaa !important; }

/* TABS */
.tabs { display: flex; border-bottom: 2px solid var(--blue); margin: 24px 0 0; }
.tab-btn {
  padding: 10px 24px; cursor: pointer; font-size: 14px; font-weight: bold;
  border: none; background: #E8EEF7; color: #555; border-radius: 6px 6px 0 0;
  margin-right: 4px; transition: background 0.15s;
}
.tab-btn.active { background: var(--blue); color: #fff; }
.tab-btn:hover:not(.active) { background: #d0d8ea; }
.tab-content { display: none; background: var(--card); border-radius: 0 0 10px 10px;
               box-shadow: 0 2px 8px rgba(0,0,0,.07); overflow-x: auto; }
.tab-content.active { display: block; }
.atab { display: none; }
.atab.active { display: block; }
.tab-section { margin-bottom: 32px; }

/* RESPONSIVE */
@media(max-width:600px){
  .hero h1 { font-size: 20px; }
  td, th   { padding: 6px 8px; font-size: 12px; }
  .pills   { gap: 10px; }
  .pill    { min-width: 90px; padding: 10px 14px; }
  .pill .val { font-size: 18px; }
  .tab-btn { padding: 8px 14px; font-size: 12px; }
}
"""

def build_html(drafts, player_analytics, num_weeks, generated_at):
    wk_hdrs = "".join(f"<th>Wk {w}</th>" for w in range(1, num_weeks + 1))
    # Aggregate stats
    my_drafts    = [d for d in drafts if MY_TEAM in d["teams"]]
    top2         = sum(1 for d in my_drafts if d["my_rank"] and d["my_rank"] <= 2)
    # Leader stats across all drafts
    leader_pts   = [d["data"].get(d["ranked"][0], {}).get("total", 0.0) for d in drafts if d["ranked"]]
    avg_leader   = sum(leader_pts) / max(len(leader_pts), 1)
    high_leader  = max(leader_pts) if leader_pts else 0.0
    low_leader   = min(leader_pts) if leader_pts else 0.0


    # ── PLAYER ANALYTICS ──
    # Build per-player stats across all drafts
    # Collect all unique players from all rosters

    # ── SEASON SUMMARY TABLE ── sorted by evilbobdole points descending
    sum_rows = ""
    for d in sorted(drafts, key=lambda d: d["my_pts"], reverse=True):
        rank     = d["my_rank"]
        my_pts   = d["my_pts"]
        # 2nd place
        second       = d["ranked"][1] if len(d["ranked"]) > 1 else d["ranked"][0] if d["ranked"] else None
        second_pts   = d["data"].get(second, {}).get("total", 0.0) if second else 0.0
        # Gap to 2nd (if I'm 1st, gap to 2nd; if I'm not 1st, gap to whoever is 2nd)
        if rank == 1:
            gap = round(my_pts - second_pts, 2)
        else:
            gap = round(my_pts - second_pts, 2)
        gap_str  = (f'<span style="color:#1a7a1a;font-weight:bold">+{gap:.2f}</span>' if gap > 0
                    else f'<span style="color:#b00;font-weight:bold">{gap:.2f}</span>' if gap < 0
                    else "0.00")
        # Pass/fail = top 2
        result   = '<span style="color:#1a7a1a;font-weight:bold">✓ PASS</span>' if rank and rank <= 2 else '<span style="color:#b00">✗</span>'
        me_cls   = "my-row" if MY_TEAM in d["teams"] else ""
        # Daily figures for this draft
        my_daily     = d.get("my_daily", 0.0)
        my_daily_gap = d.get("my_daily_gap", 0.0)
        dgap_color   = "#1a7a1a" if my_daily_gap >= 0 else "#b00"
        dgap_sign    = "+" if my_daily_gap >= 0 else ""
        sum_rows += f"""
        <tr class="{me_cls}">
          <td><a href="#d{d['num']}">Draft {d['num']}</a></td>
          <td class="num">{my_pts:.2f}</td>
          <td class="rank-cell">{ordinal(rank)}</td>
          <td class="num">{second_pts:.2f}</td>
          <td class="num">{gap_str}</td>
          <td style="text-align:center">{result}</td>
          <td class="num">{my_daily:.2f}</td>
          <td class="num" style="color:{dgap_color};font-weight:bold">{dgap_sign}{my_daily_gap:.2f}</td>
          <td class="num">{d.get("my_yesterday", 0.0):.2f}</td>
          <td class="num" style="color:{"#1a7a1a" if d.get("my_yesterday_gap",0)>=0 else "#b00"};font-weight:bold">{("+" if d.get("my_yesterday_gap",0)>=0 else "")}{d.get("my_yesterday_gap",0.0):.2f}</td>
        </tr>"""

    summary = f"""
    <div id="summary" class="summary">
      <div class="sec-title">📊 Season Summary — {MY_TEAM}</div>
      <table>
        <thead>
          <tr><th>Draft</th><th>My Points</th><th>Rank</th>
              <th>2nd Place Pts</th><th>Gap to 2nd</th><th>Result</th>
              <th>Today</th><th>Today vs 2nd</th>
              <th>Yesterday</th><th>Yest vs 2nd</th></tr>
        </thead>
        <tbody>{sum_rows}</tbody>
      </table>
    </div>"""

    # ── PER-DRAFT CARDS ──
    draft_cards = ""
    for d in drafts:
        # Leaderboard rows
        ldr_rows = ""
        for rank, team in enumerate(d["ranked"], 1):
            sc      = d["data"].get(team, {})
            total   = sc.get("total", 0.0)
            weeks   = "".join(f'<td class="num">{w:.2f}</td>' for w in sc.get("weeks", []))
            r_cls   = f"r{rank}" if rank in (1,2,3) else ""
            me_cls  = "my-row" if team == MY_TEAM else ""
            ldr_rows += f"""
            <tr class="{me_cls}">
              <td class="rank-cell {r_cls}">{rank}</td>
              <td class="bold">{team}</td>
              <td class="num bold">{total:.2f}</td>
              {weeks}
            </tr>"""

        # Player section (evilbobdole only)
        player_html = ""
        if MY_TEAM in d["teams"] and d["my_players"]:
            # My players table
            p_rows = ""
            for p in d["my_players"]:
                wkd   = "".join(f'<td class="num">{w:.2f}</td>' for w in p["weeks"])
                daily = p.get("daily", 0.0)
                yest  = p.get("yesterday", 0.0)
                p_rows += f"""
                <tr class="pos-{p['pos']}">
                  <td>{p['name']}</td>
                  <td>{p['pos']}</td>
                  <td>{p['mlb']}</td>
                  <td class="num bold">{p['total']:.2f}</td>
                  {wkd}
                  <td class="num" style="color:#888">{yest:.2f}</td>
                  <td class="num" style="color:#1a7a1a;font-weight:bold">{daily:.2f}</td>
                </tr>"""

            # 2nd place players table
            second_team = d.get("second_team", "")
            s_rows = ""
            for p in d.get("second_players", []):
                wkd   = "".join(f'<td class="num">{w:.2f}</td>' for w in p["weeks"])
                daily = p.get("daily", 0.0)
                syest = p.get("yesterday", 0.0)
                s_rows += f"""
                <tr class="pos-{p['pos']}">
                  <td>{p['name']}</td>
                  <td>{p['pos']}</td>
                  <td>{p['mlb']}</td>
                  <td class="num bold">{p['total']:.2f}</td>
                  {wkd}
                  <td class="num" style="color:#888">{syest:.2f}</td>
                  <td class="num" style="color:#cc6600;font-weight:bold">{daily:.2f}</td>
                </tr>"""

            # Bench rows
            bench_rows = ""
            _my_bench = d.get("my_bench", [])
            print(f"  Draft {d['num']}: my_bench has {len(_my_bench)} players, second_bench has {len(d.get('second_bench',[]))} players")
            for p in _my_bench:
                wkd   = "".join(f'<td class="num">{w:.2f}</td>' for w in p["weeks"])
                bench_rows += f"""
                <tr class="pos-{p['pos']}">
                  <td style="color:#666">{p['name']}</td>
                  <td>{p['pos']}</td>
                  <td>{p['mlb']}</td>
                  <td class="num">{p['total']:.2f}</td>
                  {wkd}
                  <td class="num" style="color:#aaa">{p.get('yesterday',0.0):.2f}</td>
                  <td class="num" style="color:#aaa">{p.get('daily',0.0):.2f}</td>
                </tr>"""

            s_bench_rows = ""
            for p in d.get("second_bench", []):
                wkd   = "".join(f'<td class="num">{w:.2f}</td>' for w in p["weeks"])
                s_bench_rows += f"""
                <tr class="pos-{p['pos']}">
                  <td style="color:#666">{p['name']}</td>
                  <td>{p['pos']}</td>
                  <td>{p['mlb']}</td>
                  <td class="num">{p['total']:.2f}</td>
                  {wkd}
                  <td class="num" style="color:#aaa">{p.get('yesterday',0.0):.2f}</td>
                  <td class="num" style="color:#aaa">{p.get('daily',0.0):.2f}</td>
                </tr>"""

            # Totals rows
            my_yest_tot  = d.get("my_yesterday", 0.0)
            my_today_tot = d.get("my_daily",     0.0)
            s_yest_tot   = d.get("second_yesterday", 0.0)
            s_today_tot  = d.get("second_today",     0.0)
            yg = d.get("my_yesterday_gap", 0.0); yg_col = "#1a7a1a" if yg>=0 else "#b00"; yg_s = "+" if yg>=0 else ""
            tg = d.get("my_daily_gap",     0.0); tg_col = "#1a7a1a" if tg>=0 else "#b00"; tg_s = "+" if tg>=0 else ""
            empty_wk = "".join("<td></td>" for _ in range(num_weeks))

            player_html = f"""
            <div class="player-section">
              <div style="display:grid;grid-template-columns:1fr 1fr;gap:0">
                <div>
                  <div class="player-header">⚾ {MY_TEAM} — Player Scores</div>
                  <table>
                    <thead><tr><th>Player</th><th>Pos</th><th>MLB</th><th>Total</th>{wk_hdrs}<th>Yest</th><th>Today</th></tr></thead>
                    <tbody>{p_rows}
                    <tr style="background:#d0f5d0;font-weight:bold;border-top:2px solid #333">
                      <td colspan="4">Totals</td>{empty_wk}
                      <td class="num">{my_yest_tot:.2f}</td>
                      <td class="num">{my_today_tot:.2f}</td>
                    </tr>
                    </tbody>
                  </table>
                </div>
                <div style="border-left:2px solid #ccc">
                  <div class="player-header" style="background:#2E75B6">🏆 2nd Place: {second_team}</div>
                  <table>
                    <thead><tr><th>Player</th><th>Pos</th><th>MLB</th><th>Total</th>{wk_hdrs}<th>Yest</th><th>Today</th></tr></thead>
                    <tbody>{s_rows}
                    <tr style="background:#dce8f5;font-weight:bold;border-top:2px solid #333">
                      <td colspan="4">Totals</td>{empty_wk}
                      <td class="num">{s_yest_tot:.2f}</td>
                      <td class="num">{s_today_tot:.2f}</td>
                    </tr>
                    </tbody>
                  </table>
                </div>
              </div>
              <div style="background:#f0f0f0;padding:8px 16px;font-size:12px;display:flex;gap:24px;border-top:2px solid #ccc">
                <span>Yesterday gap: <strong style="color:{yg_col}">{yg_s}{yg:.2f}</strong></span>
                <span>Today gap: <strong style="color:{tg_col}">{tg_s}{tg:.2f}</strong></span>
              </div>
              <div style="display:grid;grid-template-columns:1fr 1fr;gap:0;border-top:3px solid #aaa">
                <div>
                  <div style="background:#888;color:#fff;padding:6px 16px;font-size:12px;font-weight:bold">
                    📋 {MY_TEAM} — Bench
                  </div>
                  <table>
                    <thead><tr><th>Player</th><th>Pos</th><th>MLB</th><th>Total</th>{wk_hdrs}<th>Yest</th><th>Today</th></tr></thead>
                    <tbody>{bench_rows}</tbody>
                  </table>
                </div>
                <div style="border-left:2px solid #ccc">
                  <div style="background:#5a7fa8;color:#fff;padding:6px 16px;font-size:12px;font-weight:bold">
                    📋 {second_team} — Bench
                  </div>
                  <table>
                    <thead><tr><th>Player</th><th>Pos</th><th>MLB</th><th>Total</th>{wk_hdrs}<th>Yest</th><th>Today</th></tr></thead>
                    <tbody>{s_bench_rows}</tbody>
                  </table>
                </div>
              </div>
            </div>"""

        if MY_TEAM in d["teams"]:
            daily_val  = d.get("my_daily", 0.0)
            daily_gap  = d.get("my_daily_gap", 0.0)
            date_lbl   = d.get("latest_date", "")
            gap_color  = "#90EE90" if daily_gap >= 0 else "#FF9999"
            gap_sign   = "+" if daily_gap >= 0 else ""
            badge = (f'<span class="badge">{MY_TEAM}: {d["my_pts"]:.2f} pts &nbsp;|&nbsp; {ordinal(d["my_rank"])}</span>'
                     f'<span style="background:#333;color:#fff;padding:3px 10px;border-radius:12px;font-size:11px;margin-left:6px;">' 
                     f'{date_lbl}: {daily_val:.2f} pts &nbsp;|&nbsp; <span style="color:{gap_color}">{gap_sign}{daily_gap:.2f} vs 2nd</span></span>')
        else:
            badge = ""
        back_link = '<a href="#summary" style="color:#aac;font-size:11px;text-decoration:none;">▲ Back to Leaderboard</a>'

        draft_cards += f"""
        <div id="d{d['num']}" class="draft-card">
          <div class="draft-header">
            <h2>Draft {d['num']}</h2>
            {badge}
            {back_link}
          </div>
          <table>
            <thead><tr><th>#</th><th>Team</th><th>Total</th>{wk_hdrs}</tr></thead>
            <tbody>{ldr_rows}</tbody>
          </table>
          {player_html}
        </div>"""

    # ── FULL PAGE ──
    # ── NAV ──
    nav_links = "".join(f'<a href="#d{d["num"]}">D{d["num"]}</a>' for d in drafts)
    nav = f'<nav><span class="brand">⚾ Best Ball $600K</span><a href="#summary">Summary</a>{nav_links}</nav>'

    pills = f"""
    <div class="pills">
      <div class="pill"><div class="val">{top2}</div><div class="lbl">Top-2 Finishes</div></div>
      <div class="pill"><div class="val">{avg_leader:.2f}</div><div class="lbl">Avg Leader Score</div></div>
      <div class="pill"><div class="val">{high_leader:.2f}</div><div class="lbl">High Leader Score</div></div>
      <div class="pill"><div class="val">{low_leader:.2f}</div><div class="lbl">Low Leader Score</div></div>
    </div>"""

    # ── PLAYER ANALYTICS TABS ──
    wk_cols = "".join(f"<th>Wk {w}</th>" for w in range(num_weeks, 0, -1))
    # ── SEASON SUMMARY TABLE (used as first analytics tab) ──
    summary_table = f"""
    <table>
      <thead>
        <tr><th>Draft</th><th>My Points</th><th>Rank</th>
            <th>2nd Place Pts</th><th>Gap to 2nd</th><th>Result</th>
            <th>Today</th><th>Daily vs 2nd</th></tr>
      </thead>
      <tbody>{sum_rows}</tbody>
    </table>"""

    # ── PLAYER ANALYTICS (4 tabs: Summary, P, IF, OF) ──
    wk_hdrs_rev = "".join(f"<th>Wk {w}</th>" for w in range(num_weeks, 0, -1))

    def analytics_table(pos_filter):
        pos_players = [p for p in player_analytics if p["pos"] == pos_filter]
        pos_players.sort(key=lambda x: x["season_total"], reverse=True)
        rows = ""
        for p in pos_players:
            wk_tds = "".join(
                f'<td class="num">{p["week_totals"][w-1] if w <= len(p["week_totals"]) else 0.0:.2f}</td>'
                for w in range(num_weeks, 0, -1)
            )
            rows += f"""<tr>
              <td>{p["name"]}</td>
              <td style="text-align:center">{p["drafted_by_me"]}</td>
              <td style="text-align:center">{p["cashing"]}</td>
              <td class="num bold">{p["season_total"]:.2f}</td>
              {wk_tds}
            </tr>"""
        return f"""
        <table>
          <thead><tr>
            <th>Player</th><th>Drafted</th><th>Cashing</th><th>Season</th>{wk_hdrs_rev}
          </tr></thead>
          <tbody>{rows}</tbody>
        </table>"""

    analytics_section = f"""
    <div class="summary" style="margin-top:24px">
      <div class="sec-title">📊 Season Overview — evilbobdole</div>
      <div style="display:flex;gap:0;border-bottom:2px solid #1F4E79;padding:0 16px;background:#f8f9fc">
        <button class="tab-btn active" onclick="showAnalyticsTab('SUM',this)">📋 Summary</button>
        <button class="tab-btn" onclick="showAnalyticsTab('P',this)">⚾ Pitchers</button>
        <button class="tab-btn" onclick="showAnalyticsTab('IF',this)">🏃 Infielders</button>
        <button class="tab-btn" onclick="showAnalyticsTab('OF',this)">🌴 Outfielders</button>
      </div>
      <div id="atab-SUM" class="atab active">{summary_table}</div>
      <div id="atab-P"   class="atab">{analytics_table("P")}</div>
      <div id="atab-IF"  class="atab">{analytics_table("IF")}</div>
      <div id="atab-OF"  class="atab">{analytics_table("OF")}</div>
    </div>"""


    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>⚾ MLB Best Ball $600K Slugfest</title>
<style>{CSS}
.tab-btn {{
  padding: 10px 20px; background: #E8EEF7; border: none; cursor: pointer;
  font-size: 13px; font-weight: bold; color: #333; border-right: 1px solid #ccc;
}}
.tab-btn.active {{ background: #1F4E79; color: #fff; }}
.tab-btn:hover:not(.active) {{ background: #d0daf0; }}
.tab-panel {{ overflow-x: auto; }}
</style>
</head>
<body>
{nav}
<div class="hero">
  <h1>⚾ MLB Best Ball $600K Slugfest</h1>
  <div class="sub">35 Drafts &nbsp;·&nbsp; DraftKings Scoring &nbsp;·&nbsp; Tracking: {MY_TEAM}</div>
  <div class="ts">Updated: {generated_at}</div>
</div>
{pills}
<div class="container">
{analytics_section}
{draft_cards}
</div>
<script>
// Tab switching
function showAnalyticsTab(pos, btn) {{
  document.querySelectorAll('.atab').forEach(t => t.classList.remove('active'));
  document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
  document.getElementById('atab-'+pos).classList.add('active');
  btn.classList.add('active');
}}

// Smooth scroll
document.querySelectorAll('a[href^="#"]').forEach(a => {{
  a.addEventListener('click', e => {{
    e.preventDefault();
    const el = document.querySelector(a.getAttribute('href'));
    if (el) el.scrollIntoView({{behavior:'smooth', block:'start'}});
  }});
}});

// Auto-refresh during live games
// Uses the MLB schedule API to check for active games.
// Stats are keyed by scheduled date, so a game starting 10pm that ends
// after midnight still counts as the scheduled day — no date confusion.
(async function checkLiveGames() {{
  const statusBadge = document.createElement('div');
  statusBadge.id = 'live-status';
  statusBadge.style.cssText = 'position:fixed;bottom:16px;right:16px;padding:6px 12px;' +
    'border-radius:20px;font-size:12px;font-weight:bold;z-index:999;display:none;';
  document.body.appendChild(statusBadge);

  function showStatus(msg, color) {{
    statusBadge.textContent = msg;
    statusBadge.style.background = color;
    statusBadge.style.color = '#fff';
    statusBadge.style.display = 'block';
  }}

  async function checkGames() {{
    try {{
      // Use today's date in ET (MLB schedule is Eastern Time)
      const now = new Date();
      // Format as YYYY-MM-DD
      const pad = n => String(n).padStart(2,'0');
      const dateStr = `${{now.getFullYear()}}-${{pad(now.getMonth()+1)}}-${{pad(now.getDate())}}`;

      const url = `https://statsapi.mlb.com/api/v1/schedule?sportId=1&date=${{dateStr}}&hydrate=game(status)`;
      const resp = await fetch(url);
      const data = await resp.json();

      let liveCount = 0;
      let totalGames = 0;

      if (data.dates && data.dates.length > 0) {{
        for (const date of data.dates) {{
          for (const game of (date.games || [])) {{
            totalGames++;
            const status = game.status?.abstractGameState;
            // "Live" = game in progress
            if (status === 'Live') liveCount++;
          }}
        }}
      }}

      if (liveCount > 0) {{
        showStatus(`⚾ ${{liveCount}} game${{liveCount>1?'s':''}} live — refreshing in 15 min`, '#e74c3c');
        // Refresh page in 15 minutes
        setTimeout(() => location.reload(), 15 * 60 * 1000);
      }} else if (totalGames > 0) {{
        showStatus(`✓ All ${{totalGames}} games final`, '#27ae60');
        // Hide after 5 seconds
        setTimeout(() => {{ statusBadge.style.display='none'; }}, 5000);
      }}
    }} catch(e) {{
      // Silently fail if API unreachable
      console.log('MLB schedule check failed:', e);
    }}
  }}

  // Check on load
  await checkGames();
}})();
</script>
</body>
</html>"""


# ── MAIN ──────────────────────────────────────────────────────────────────────

def git_push(repo_dir, message):
    # Stage and commit index.html
    for cmd in [
        ["git", "-C", str(repo_dir), "add", "index.html"],
        ["git", "-C", str(repo_dir), "commit", "-m", message],
    ]:
        r = subprocess.run(cmd, capture_output=True, text=True)
        combined = r.stdout + r.stderr
        if r.returncode != 0:
            if "nothing to commit" in combined:
                print("No changes since last push.")
                return
            print(f"stdout: {r.stdout}")
            print(f"stderr: {r.stderr}")
            raise RuntimeError(f"Git command failed: {' '.join(cmd)}")

    # Fetch latest, reset local to match remote, then push our index.html on top
    subprocess.run(["git", "-C", str(repo_dir), "fetch", "origin"], capture_output=True)
    subprocess.run(["git", "-C", str(repo_dir), "reset", "--soft", "origin/main"], capture_output=True)
    # Re-add index.html after reset in case it got unstaged
    subprocess.run(["git", "-C", str(repo_dir), "add", "index.html"], capture_output=True)
    r = subprocess.run(
        ["git", "-C", str(repo_dir), "push"],
        capture_output=True, text=True
    )
    if r.returncode != 0:
        print(f"stdout: {r.stdout}")
        print(f"stderr: {r.stderr}")
        raise RuntimeError("Git push failed")
    print("✓ Pushed to GitHub Pages")

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--local",  action="store_true", help="Skip git push")
    parser.add_argument("--xlsx",   default=str(XLSX_PATH))
    parser.add_argument("--weeks",  type=int, default=1)
    args = parser.parse_args()

    xlsx = Path(args.xlsx)
    if not xlsx.exists():
        raise FileNotFoundError(f"Not found: {xlsx}")

    print(f"Loading {xlsx.name}...")
    wb = load_workbook(xlsx, data_only=True)

    print("Reading stats...")
    batting, pitching, batting_daily, pitching_daily = load_stats(wb)
    print(f"  Batting rows:  {len(batting)}")
    print(f"  Pitching rows: {len(pitching)}")

    # Find the two most recent dates in stats
    all_dates = sorted(set(k[0] for k in batting_daily) | set(k[0] for k in pitching_daily))
    latest_date    = all_dates[-1] if len(all_dates) >= 1 else None
    yesterday_date = all_dates[-2] if len(all_dates) >= 2 else None
    print(f"  Latest date:   {latest_date}")
    print(f"  Yesterday:     {yesterday_date}")

    # Build player list for analytics from draft rosters (no ADP file needed)
    # Collect all unique players across all drafts with pos/mlb info
    adp_players = []
    seen_players = set()
    try:
        wb_check = load_workbook(xlsx, data_only=True)
        for sheet_name in wb_check.sheetnames:
            if not sheet_name.startswith("draftboard_"): continue
            ws = wb_check[sheet_name]
            for row in range(2, 242):
                name = ws.cell(row, 2).value
                pos  = ws.cell(row, 3).value
                mlb  = ws.cell(row, 4).value
                if not name or not pos: continue
                name_s = strip_accents(str(name).strip())
                pos_s  = str(pos).strip()
                mlb_s  = str(mlb).strip().upper().replace("AZ","ARI") if mlb else ""
                key    = (name_s, pos_s)
                if key not in seen_players and pos_s in ("P","IF","OF"):
                    seen_players.add(key)
                    adp_players.append({"name": name_s, "pos": pos_s, "mlb": mlb_s})
        wb_check.close()
        print(f"  Unique players for analytics: {len(adp_players)}")
    except Exception as e:
        print(f"  Player list build failed: {e}")

    result = load_all_drafts(wb, batting, pitching, batting_daily, pitching_daily,
                             latest_date, yesterday_date, num_weeks=args.weeks, adp_players=adp_players)
    drafts, player_analytics = result

    generated_at = datetime.now().strftime("%B %d, %Y at %I:%M %p")
    print("Building HTML...")
    html = build_html(drafts, player_analytics, args.weeks, generated_at)

    out = Path("index.html") if args.local else REPO_DIR / "index.html"
    out.write_text(html, encoding="utf-8")
    print(f"✓ Written: {out}  ({len(html):,} bytes)")

    if not args.local:
        print("Pushing to GitHub...")
        git_push(REPO_DIR, f"Update scores {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        username = subprocess.run(
            ["git", "-C", str(REPO_DIR), "remote", "get-url", "origin"],
            capture_output=True, text=True
        ).stdout.strip()
        repo_name = Path(username).stem
        print(f"\n✓ Live at: https://YOUR_USERNAME.github.io/{repo_name}/")
    else:
        print(f"\n✓ Open index.html in your browser to preview.")

if __name__ == "__main__":
    main()
